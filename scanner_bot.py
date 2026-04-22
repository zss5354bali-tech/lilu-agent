"""
Scanner Bot — база поставщиков (чистая копия suppliers_bot без исторических данных).
Принимает фото визиток/буклетов и голосовые комментарии,
извлекает данные через Claude Vision + Whisper,
сохраняет в SQLite, экспортирует в красивый Excel.

Env vars:
  SCANNER_BOT_TOKEN   — токен Telegram бота
  ANTHROPIC_API_KEY   — Claude API
  OPENAI_API_KEY      — Whisper API
"""

import logging
import os
import asyncio
import json
import re
import sqlite3
import datetime
import io
import base64

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, filters, ContextTypes,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN         = os.getenv("SCANNER_BOT_TOKEN", "8692863987:AAFBRaEG9rwNcynoZsAZrtPN1ksKxhQD2eg")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
OPENAI_API_KEY    = os.getenv("OPENAI_API_KEY")

_DATA_DIR = "/data" if os.path.isdir("/data") else "."
DB_PATH   = os.path.join(_DATA_DIR, "scanner.db")


# ─── Database ───────────────────────────────────────────────────────────────

def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS suppliers (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            company       TEXT,
            contact_name  TEXT,
            phone         TEXT,
            email         TEXT,
            website       TEXT,
            products      TEXT,
            country       TEXT,
            city          TEXT,
            voice_comment TEXT,
            created_at    TEXT
        )
    """)
    conn.commit()
    conn.close()


def save_supplier(data: dict) -> int:
    conn = sqlite3.connect(DB_PATH)
    cur = conn.execute(
        """INSERT INTO suppliers
           (company, contact_name, phone, email, website, products,
            country, city, voice_comment, created_at)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (
            data.get("company", ""),
            data.get("contact_name", ""),
            data.get("phone", ""),
            data.get("email", ""),
            data.get("website", ""),
            data.get("products", ""),
            data.get("country", ""),
            data.get("city", ""),
            data.get("voice_comment", ""),
            datetime.datetime.now().isoformat(),
        ),
    )
    row_id = cur.lastrowid
    conn.commit()
    conn.close()
    return row_id


def get_all_suppliers() -> list[dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT * FROM suppliers ORDER BY created_at DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def search_suppliers(query: str) -> list[dict]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    words = query.lower().split()
    if not words:
        return []
    conditions = []
    params = []
    for word in words:
        conditions.append(
            "(lower(company) LIKE ? OR lower(contact_name) LIKE ? OR "
            "lower(products) LIKE ? OR lower(email) LIKE ? OR "
            "lower(phone) LIKE ? OR lower(voice_comment) LIKE ?)"
        )
        p = f"%{word}%"
        params.extend([p, p, p, p, p, p])
    where = " AND ".join(conditions)
    rows = conn.execute(
        f"SELECT * FROM suppliers WHERE {where} ORDER BY created_at DESC", params
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ─── AI ─────────────────────────────────────────────────────────────────────

async def transcribe_voice(file_bytes: bytes) -> str:
    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.post(
            "https://api.openai.com/v1/audio/transcriptions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
            files={"file": ("voice.ogg", file_bytes, "audio/ogg")},
            data={"model": "whisper-1"},
        )
        response.raise_for_status()
        return response.json()["text"]


def _parse_json_robust(text: str) -> dict:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    clean = re.sub(r'```(?:json)?\s*', '', text).strip().rstrip('`').strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        pass

    start = text.find('{')
    end = text.rfind('}')
    if start != -1 and end > start:
        try:
            return json.loads(text[start:end + 1])
        except json.JSONDecodeError:
            pass

    fields = ["company", "contact_name", "phone", "email", "website",
              "products", "country", "city", "voice_comment"]
    result = {f: "" for f in fields}
    for field in fields:
        m = re.search(rf'"{field}"\s*:\s*"((?:[^"\\]|\\.)*)"', text)
        if m:
            result[field] = m.group(1)
    return result


async def extract_supplier_data(photos_b64: list[str], voice_text: str) -> dict:
    content = []
    for b64 in photos_b64:
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": b64},
        })

    prompt = "Извлеки информацию о поставщике"
    if photos_b64:
        prompt += " из изображений (визитки, буклеты, прайс-листы)"
    if voice_text:
        prompt += f"\n\nГолосовой комментарий пользователя:\n{voice_text}"

    prompt += """

Верни ТОЛЬКО валидный JSON — без markdown, без пояснений, без переносов строк внутри значений.
Все значения — однострочные строки, спецсимволы экранировать.

{"company":"","contact_name":"","phone":"","email":"","website":"","products":"","country":"","city":"","voice_comment":""}

Заполни поля по данным. Если поле не найдено — пустая строка."""

    content.append({"type": "text", "text": prompt})

    async with httpx.AsyncClient(timeout=60) as client:
        response = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-6",
                "max_tokens": 1024,
                "messages": [{"role": "user", "content": content}],
            },
        )
        response.raise_for_status()
        text = response.json()["content"][0]["text"].strip()

    return _parse_json_robust(text)


# ─── State ──────────────────────────────────────────────────────────────────

user_states: dict[int, dict] = {}


def get_state(user_id: int) -> dict:
    if user_id not in user_states:
        user_states[user_id] = {"mode": "idle", "photos": [], "voices": [], "extracted": {}}
    return user_states[user_id]


def reset_state(user_id: int):
    user_states[user_id] = {"mode": "idle", "photos": [], "voices": [], "extracted": {}}


# ─── Keyboards ──────────────────────────────────────────────────────────────

def main_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 Новая запись", callback_data="new_entry")],
        [
            InlineKeyboardButton("📋 Список", callback_data="list"),
            InlineKeyboardButton("🔍 Поиск", callback_data="search"),
        ],
        [InlineKeyboardButton("📥 Экспорт Excel", callback_data="export")],
    ])


def collecting_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Готово — обработать", callback_data="done")],
        [InlineKeyboardButton("❌ Отмена", callback_data="cancel")],
    ])


def confirm_kb():
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("💾 Сохранить", callback_data="save"),
            InlineKeyboardButton("❌ Отмена", callback_data="cancel"),
        ],
    ])


# ─── Formatting ─────────────────────────────────────────────────────────────

def esc(text: str) -> str:
    return (text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def fmt_card(data: dict, sid: int = None) -> str:
    lines = []
    if sid:
        lines.append(f"<b>#{sid}</b>")
    if data.get("company"):
        lines.append(f"🏢 <b>{esc(data['company'])}</b>")
    if data.get("contact_name"):
        lines.append(f"👤 {esc(data['contact_name'])}")
    if data.get("phone"):
        lines.append(f"📞 {esc(data['phone'])}")
    if data.get("email"):
        lines.append(f"📧 {esc(data['email'])}")
    if data.get("website"):
        lines.append(f"🌐 {esc(data['website'])}")
    if data.get("products"):
        lines.append(f"📦 {esc(data['products'])}")
    location = ", ".join(filter(None, [data.get("city"), data.get("country")]))
    if location:
        lines.append(f"📍 {esc(location)}")
    if data.get("voice_comment"):
        lines.append(f"💬 <i>{esc(data['voice_comment'])}</i>")
    return "\n".join(lines) if lines else "<i>(данные не извлечены)</i>"


# ─── Handlers ───────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reset_state(update.effective_user.id)
    await update.message.reply_text(
        "👋 <b>База поставщиков</b>\n\n"
        "Нажмите <b>Новая запись</b>, затем отправьте:\n"
        "• 📷 фото визитки или буклета\n"
        "• 🎤 голосовой комментарий\n\n"
        "Можно несколько фото и голосовых. Когда всё — нажмите <b>Готово</b>.",
        parse_mode="HTML",
        reply_markup=main_kb(),
    )


async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    state = get_state(user_id)
    data = query.data

    # ── Новая запись ──
    if data == "new_entry":
        reset_state(user_id)
        state = get_state(user_id)
        state["mode"] = "collecting"
        await query.message.reply_text(
            "📝 <b>Новая запись</b>\n\n"
            "Отправляйте фото и голосовые — в любом порядке и количестве.\n"
            "Когда закончите — нажмите <b>Готово</b>.",
            parse_mode="HTML",
            reply_markup=collecting_kb(),
        )

    # ── Готово — обработать ──
    elif data == "done":
        if state["mode"] != "collecting":
            await query.message.reply_text("Сначала нажмите «Новая запись».")
            return
        if not state["photos"] and not state["voices"]:
            await query.message.reply_text(
                "Нет данных. Отправьте фото или голосовое сообщение.",
                reply_markup=collecting_kb(),
            )
            return

        state["mode"] = "processing"
        msg = await query.message.reply_text("⏳ Обрабатываю данные...")
        try:
            voice_text = "\n".join(state["voices"])
            extracted = await extract_supplier_data(state["photos"], voice_text)
            state["extracted"] = extracted
            state["mode"] = "confirming"
            await msg.edit_text(
                f"Проверьте данные:\n\n{fmt_card(extracted)}\n\nСохранить в базу?",
                parse_mode="HTML",
                reply_markup=confirm_kb(),
            )
        except Exception as e:
            logger.error(f"Extraction error: {e}")
            state["mode"] = "collecting"
            await msg.edit_text(
                f"⚠️ Ошибка обработки: {e}\n\nПопробуйте ещё раз.",
                reply_markup=collecting_kb(),
            )

    # ── Сохранить ──
    elif data == "save":
        if state["mode"] != "confirming" or not state["extracted"]:
            return
        extracted = state["extracted"].copy()
        supplier_id = save_supplier(extracted)
        reset_state(user_id)
        await query.message.reply_text(
            f"✅ Сохранено!\n\n{fmt_card(extracted, supplier_id)}",
            parse_mode="HTML",
            reply_markup=main_kb(),
        )

    # ── Отмена ──
    elif data == "cancel":
        reset_state(user_id)
        await query.message.reply_text("Отменено.", reply_markup=main_kb())

    # ── Список ──
    elif data == "list":
        suppliers = get_all_suppliers()
        if not suppliers:
            await query.message.reply_text("База пуста.", reply_markup=main_kb())
            return
        text = f"📋 <b>Поставщики — {len(suppliers)} шт.:</b>\n\n"
        for s in suppliers[:20]:
            name = esc(s.get("company") or s.get("contact_name") or "—")
            text += f"• <b>#{s['id']}</b> {name}"
            if s.get("products"):
                text += f" — {esc(s['products'][:50])}"
            text += "\n"
        if len(suppliers) > 20:
            text += f"\n<i>...и ещё {len(suppliers) - 20}. Используйте экспорт для полного списка.</i>"
        await query.message.reply_text(text, parse_mode="HTML", reply_markup=main_kb())

    # ── Поиск ──
    elif data == "search":
        state["mode"] = "searching"
        await query.message.reply_text("🔍 Введите запрос для поиска:")

    # ── Экспорт ──
    elif data == "export":
        suppliers = get_all_suppliers()
        if not suppliers:
            await query.message.reply_text("База пуста.", reply_markup=main_kb())
            return

        fields  = ["id", "company", "contact_name", "phone", "email", "website",
                   "products", "country", "city", "voice_comment", "created_at"]
        headers = ["ID", "Компания", "Контакт", "Телефон", "Email", "Сайт",
                   "Продукты", "Страна", "Город", "Голосовой комментарий", "Дата"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Поставщики"

        header_fill  = PatternFill("solid", fgColor="2E4057")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin         = Side(style="thin", color="CCCCCC")
        cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        col_widths   = [6, 28, 22, 18, 28, 28, 40, 16, 16, 40, 20]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 30

        fill_even       = PatternFill("solid", fgColor="F0F4F8")
        fill_odd        = PatternFill("solid", fgColor="FFFFFF")
        align_wrap      = Alignment(vertical="top", wrap_text=True)
        align_nowrap    = Alignment(vertical="top", wrap_text=False)

        for row_idx, s in enumerate(suppliers, start=2):
            fill = fill_even if row_idx % 2 == 0 else fill_odd
            for col_idx, field in enumerate(fields, start=1):
                value = s.get(field) or ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill   = fill
                cell.border = cell_border
                cell.alignment = align_wrap if field in ("products", "voice_comment") else align_nowrap

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"suppliers_{datetime.date.today()}.xlsx"
        await query.message.reply_document(
            document=InputFile(buf, filename=filename),
            caption=f"📥 База поставщиков — {len(suppliers)} записей",
            reply_markup=main_kb(),
        )


async def on_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state = get_state(user_id)
    if state["mode"] != "collecting":
        await update.message.reply_text(
            "Нажмите «Новая запись» чтобы начать.", reply_markup=main_kb()
        )
        return

    photo = update.message.photo[-1]
    file  = await context.bot.get_file(photo.file_id)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(file.file_path)
    b64 = base64.b64encode(resp.content).decode()
    state["photos"].append(b64)

    count_p = len(state["photos"])
    count_v = len(state["voices"])
    status  = f"📷 Фото: {count_p}"
    if count_v:
        status += f"  🎤 Голосовых: {count_v}"
    await update.message.reply_text(status, reply_markup=collecting_kb())


async def on_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state   = get_state(user_id)
    if state["mode"] != "collecting":
        await update.message.reply_text(
            "Нажмите «Новая запись» чтобы начать.", reply_markup=main_kb()
        )
        return

    msg  = await update.message.reply_text("🎤 Распознаю голос...")
    file = await context.bot.get_file(update.message.voice.file_id)
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(file.file_path)
    try:
        text = await transcribe_voice(resp.content)
        state["voices"].append(text)
        count_p = len(state["photos"])
        count_v = len(state["voices"])
        status  = f"🎤 Голосовых: {count_v}"
        if count_p:
            status += f"  📷 Фото: {count_p}"
        await msg.edit_text(
            f"{status}\n\n<i>{esc(text)}</i>", parse_mode="HTML", reply_markup=collecting_kb()
        )
    except Exception as e:
        await msg.edit_text(f"⚠️ Ошибка распознавания: {e}", reply_markup=collecting_kb())


async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    state   = get_state(user_id)

    if state["mode"] == "searching":
        state["mode"]  = "idle"
        query_text     = update.message.text.strip()
        results        = search_suppliers(query_text)
        if not results:
            await update.message.reply_text(
                f"Ничего не найдено по запросу «{query_text}».", reply_markup=main_kb()
            )
            return
        text = f"🔍 <b>Найдено: {len(results)}</b>\n\n"
        for s in results[:10]:
            text += fmt_card(s, s["id"]) + "\n\n─────\n\n"
        await update.message.reply_text(text, parse_mode="HTML", reply_markup=main_kb())
    else:
        await update.message.reply_text("Используйте кнопки ниже.", reply_markup=main_kb())


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    init_db()
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CallbackQueryHandler(on_callback))
    app.add_handler(MessageHandler(filters.PHOTO, on_photo))
    app.add_handler(MessageHandler(filters.VOICE, on_voice))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    logger.info("Scanner bot started")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
